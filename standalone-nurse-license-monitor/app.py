from __future__ import annotations

import csv
import json
import os
import smtplib
import ssl
import threading
import time
from dataclasses import dataclass
from datetime import date, datetime, timedelta
from email.message import EmailMessage
from pathlib import Path
from tkinter import BooleanVar, StringVar, Tk, filedialog, messagebox
from tkinter import ttk

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


APP_DIR = Path(__file__).resolve().parent
SETTINGS_FILE = APP_DIR / "settings.json"
HISTORY_FILE = APP_DIR / "send_history.json"
LOG_FILE = APP_DIR / "monitor.log"


DEFAULT_SETTINGS = {
    "excel_path": "",
    "sheet_name": "",
    "hr_copy_emails": "",
    "smtp_host": "",
    "smtp_port": "587",
    "smtp_username": "",
    "smtp_password": "",
    "smtp_from": "",
    "smtp_use_tls": True,
    "subject_prefix": "Credential expiration notice",
    "monitor_enabled": False,
}


ALIASES = {
    "name": [
        "nurse name",
        "nurse",
        "employee name",
        "full name",
        "name",
        "caregiver name",
        "staff name",
    ],
    "first_name": ["first name", "firstname"],
    "last_name": ["last name", "lastname"],
    "email": ["email", "email address", "nurse email", "employee email", "staff email"],
    "sms_email": [
        "email to sms",
        "sms email",
        "text email",
        "sms gateway",
        "sms gateway email",
        "text message email",
    ],
    "document": [
        "document",
        "document type",
        "document name",
        "license",
        "license type",
        "credential",
        "credential type",
        "file",
        "file type",
    ],
    "expires": [
        "expiration date",
        "expires",
        "expires at",
        "expiry",
        "expiry date",
        "expiration",
        "renewal date",
        "due date",
    ],
}


BUCKET_RULES = {
    "expired": {"label": "Expired", "limit": 3, "days": 1},
    "lt15": {"label": "Less than 15 days", "limit": 2, "days": 1},
    "lt30": {"label": "Less than 30 days", "limit": 1, "days": 1},
    "lt60": {"label": "Less than 60 days", "limit": 2, "days": 7},
    "lt90": {"label": "Less than 90 days", "limit": 3, "days": 30},
}


@dataclass
class CredentialAlert:
    nurse_name: str
    email: str
    sms_email: str
    document_name: str
    expires_on: date
    row_number: int
    days_until: int
    bucket: str
    due_now: bool = False


def log(message: str) -> None:
    stamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    with LOG_FILE.open("a", encoding="utf-8") as handle:
        handle.write(f"[{stamp}] {message}\n")


def load_settings() -> dict:
    if not SETTINGS_FILE.exists():
        return dict(DEFAULT_SETTINGS)
    try:
        with SETTINGS_FILE.open("r", encoding="utf-8") as handle:
            settings = json.load(handle)
        return {**DEFAULT_SETTINGS, **settings}
    except Exception as exc:
        log(f"Could not read settings: {exc}")
        return dict(DEFAULT_SETTINGS)


def save_settings(settings: dict) -> None:
    with SETTINGS_FILE.open("w", encoding="utf-8") as handle:
        json.dump(settings, handle, indent=2)


def load_history() -> dict:
    if not HISTORY_FILE.exists():
        return {}
    try:
        with HISTORY_FILE.open("r", encoding="utf-8") as handle:
            return json.load(handle)
    except Exception as exc:
        log(f"Could not read send history: {exc}")
        return {}


def save_history(history: dict) -> None:
    with HISTORY_FILE.open("w", encoding="utf-8") as handle:
        json.dump(history, handle, indent=2)


def normalize_header(value: object) -> str:
    return "".join(ch for ch in str(value or "").lower() if ch.isalnum())


def parse_date(value: object) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = str(value or "").strip()
    if not text:
        return None
    formats = [
        "%Y-%m-%d",
        "%m/%d/%Y",
        "%m-%d-%Y",
        "%m/%d/%y",
        "%m-%d-%y",
        "%B %d, %Y",
        "%b %d, %Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            pass
    try:
        return datetime.fromisoformat(text).date()
    except ValueError:
        return None


def split_emails(value: str) -> list[str]:
    emails: list[str] = []
    for chunk in value.replace(";", ",").replace("\n", ",").split(","):
        clean = chunk.strip()
        if clean:
            emails.append(clean)
    return emails


def find_column(headers: list[str], field: str) -> int | None:
    wanted = {normalize_header(alias) for alias in ALIASES[field]}
    for index, header in enumerate(headers):
        if normalize_header(header) in wanted:
            return index
    return None


def get_cell(row: list[object], index: int | None) -> object:
    if index is None or index >= len(row):
        return ""
    return row[index]


def bucket_for(days_until: int) -> str | None:
    if days_until < 0:
        return "expired"
    if days_until < 15:
        return "lt15"
    if days_until < 30:
        return "lt30"
    if days_until < 60:
        return "lt60"
    if days_until < 90:
        return "lt90"
    return None


def history_key(alert: CredentialAlert) -> str:
    parts = [alert.nurse_name, alert.document_name, alert.expires_on.isoformat(), alert.bucket]
    return "|".join(part.strip().lower() for part in parts)


def due_by_history(alert: CredentialAlert, history: dict, now: datetime) -> bool:
    rule = BUCKET_RULES[alert.bucket]
    cutoff = now - timedelta(days=rule["days"])
    stamps = []
    for value in history.get(history_key(alert), []):
        try:
            parsed = datetime.fromisoformat(value)
        except ValueError:
            continue
        if parsed >= cutoff:
            stamps.append(parsed)
    return len(stamps) < rule["limit"]


def rows_from_xlsx(path: Path, sheet_name: str) -> list[list[object]]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Run Start App.cmd again.")
    workbook = load_workbook(path, data_only=True, read_only=True)
    sheet = workbook[sheet_name] if sheet_name and sheet_name in workbook.sheetnames else workbook[workbook.sheetnames[0]]
    return [list(row) for row in sheet.iter_rows(values_only=True)]


def rows_from_csv(path: Path) -> list[list[object]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [list(row) for row in csv.reader(handle)]


def read_alerts(settings: dict) -> tuple[list[CredentialAlert], list[str]]:
    excel_path = Path(settings.get("excel_path", "")).expanduser()
    if not excel_path.exists():
        return [], ["Excel file was not found."]

    if excel_path.suffix.lower() == ".csv":
        rows = rows_from_csv(excel_path)
    else:
        rows = rows_from_xlsx(excel_path, settings.get("sheet_name", "").strip())

    if not rows:
        return [], ["The workbook is empty."]

    headers = [str(value or "").strip() for value in rows[0]]
    indexes = {
        "name": find_column(headers, "name"),
        "first_name": find_column(headers, "first_name"),
        "last_name": find_column(headers, "last_name"),
        "email": find_column(headers, "email"),
        "sms_email": find_column(headers, "sms_email"),
        "document": find_column(headers, "document"),
        "expires": find_column(headers, "expires"),
    }

    warnings: list[str] = []
    alerts: list[CredentialAlert] = []
    today = date.today()

    if indexes["expires"] is None:
        return [], ["No expiration date column was found."]

    for row_number, row in enumerate(rows[1:], start=2):
        if not any(str(value or "").strip() for value in row):
            continue

        first = str(get_cell(row, indexes["first_name"]) or "").strip()
        last = str(get_cell(row, indexes["last_name"]) or "").strip()
        name = str(get_cell(row, indexes["name"]) or "").strip() or " ".join(part for part in [first, last] if part)
        email = str(get_cell(row, indexes["email"]) or "").strip()
        sms_email = str(get_cell(row, indexes["sms_email"]) or "").strip()
        document = str(get_cell(row, indexes["document"]) or "").strip() or "Credential"
        expires_on = parse_date(get_cell(row, indexes["expires"]))

        if not name or not expires_on:
            warnings.append(f"Skipped row {row_number}: missing nurse name or expiration date.")
            continue

        days_until = (expires_on - today).days
        bucket = bucket_for(days_until)
        if bucket is None:
            continue

        alerts.append(
            CredentialAlert(
                nurse_name=name,
                email=email,
                sms_email=sms_email,
                document_name=document,
                expires_on=expires_on,
                row_number=row_number,
                days_until=days_until,
                bucket=bucket,
            )
        )

    history = load_history()
    now = datetime.now()
    for alert in alerts:
        alert.due_now = due_by_history(alert, history, now)

    return alerts, warnings


def build_message(alert: CredentialAlert) -> tuple[str, str]:
    if alert.days_until < 0:
        timing = f"expired on {alert.expires_on.isoformat()}"
    else:
        day_word = "day" if alert.days_until == 1 else "days"
        timing = f"will expire in {alert.days_until} {day_word} on {alert.expires_on.isoformat()}"

    subject = f"{alert.nurse_name} - {alert.document_name}"
    body = "\n".join(
        [
            f"Hello {alert.nurse_name},",
            "",
            f"Our records show that your {alert.document_name} {timing}.",
            "Please send your updated document to HR as soon as possible.",
            "",
            "Quality One Care HR",
        ]
    )
    return subject, body


def send_email(settings: dict, to_email: str, subject: str, body: str) -> None:
    host = settings.get("smtp_host", "").strip()
    port = int(settings.get("smtp_port", "587") or "587")
    username = settings.get("smtp_username", "").strip()
    password = settings.get("smtp_password", "")
    from_email = settings.get("smtp_from", "").strip() or username

    if not host or not from_email:
        raise RuntimeError("SMTP host and From Email are required before sending.")

    message = EmailMessage()
    message["From"] = from_email
    message["To"] = to_email
    message["Subject"] = subject
    message.set_content(body)

    if settings.get("smtp_use_tls", True):
        with smtplib.SMTP(host, port, timeout=30) as server:
            server.ehlo()
            server.starttls(context=ssl.create_default_context())
            server.ehlo()
            if username:
                server.login(username, password)
            server.send_message(message)
    else:
        with smtplib.SMTP_SSL(host, port, timeout=30) as server:
            if username:
                server.login(username, password)
            server.send_message(message)


def send_due_messages(settings: dict, force: bool = False) -> dict:
    alerts, warnings = read_alerts(settings)
    history = load_history()
    now = datetime.now()
    sent = 0
    skipped = 0

    for alert in alerts:
        if not force and not due_by_history(alert, history, now):
            skipped += 1
            continue

        recipients = []
        for email in [alert.email, alert.sms_email, *split_emails(settings.get("hr_copy_emails", ""))]:
            if email and email not in recipients:
                recipients.append(email)

        if not recipients:
            warnings.append(f"Skipped row {alert.row_number}: no email or email-to-SMS address.")
            skipped += 1
            continue

        subject, body = build_message(alert)
        subject_prefix = settings.get("subject_prefix", DEFAULT_SETTINGS["subject_prefix"]).strip()
        full_subject = f"{subject_prefix}: {subject}"

        for recipient in recipients:
            send_email(settings, recipient, full_subject, body)
            sent += 1

        key = history_key(alert)
        history[key] = [*history.get(key, []), now.isoformat()][-50:]

    save_history(history)
    return {"scanned": len(alerts), "sent": sent, "skipped": skipped, "warnings": warnings}


class NurseLicenseMonitorApp:
    def __init__(self) -> None:
        self.root = Tk()
        self.root.title("Nurse License Monitor")
        self.root.geometry("1060x720")
        self.root.minsize(920, 600)

        self.settings = load_settings()
        self.monitor_running = False
        self.worker_running = False

        self.vars = {key: StringVar(value=str(value)) for key, value in self.settings.items() if key != "smtp_use_tls" and key != "monitor_enabled"}
        self.smtp_use_tls = BooleanVar(value=bool(self.settings.get("smtp_use_tls", True)))
        self.monitor_enabled = BooleanVar(value=bool(self.settings.get("monitor_enabled", False)))

        self.status_text = StringVar(value="Ready")
        self.summary_text = StringVar(value="Choose an Excel file, save settings, then scan.")

        self.build_ui()
        self.refresh_preview()
        if self.monitor_enabled.get():
            self.start_monitor()

    def build_ui(self) -> None:
        style = ttk.Style()
        style.theme_use("clam")
        style.configure("TButton", padding=(10, 6))
        style.configure("Header.TLabel", font=("Segoe UI", 16, "bold"))
        style.configure("Summary.TLabel", font=("Segoe UI", 10))

        main = ttk.Frame(self.root, padding=16)
        main.pack(fill="both", expand=True)

        header = ttk.Frame(main)
        header.pack(fill="x")
        ttk.Label(header, text="Nurse License Monitor", style="Header.TLabel").pack(side="left")
        ttk.Label(header, textvariable=self.status_text).pack(side="right")

        notebook = ttk.Notebook(main)
        notebook.pack(fill="both", expand=True, pady=(14, 0))

        monitor_tab = ttk.Frame(notebook, padding=12)
        settings_tab = ttk.Frame(notebook, padding=12)
        notebook.add(monitor_tab, text="Monitor")
        notebook.add(settings_tab, text="Settings")

        self.build_monitor_tab(monitor_tab)
        self.build_settings_tab(settings_tab)

    def build_monitor_tab(self, parent: ttk.Frame) -> None:
        controls = ttk.Frame(parent)
        controls.pack(fill="x")
        ttk.Button(controls, text="Scan Now", command=self.refresh_preview).pack(side="left")
        ttk.Button(controls, text="Send Due Now", command=lambda: self.run_send(False)).pack(side="left", padx=(8, 0))
        ttk.Button(controls, text="Send All Visible", command=lambda: self.run_send(True)).pack(side="left", padx=(8, 0))
        ttk.Button(controls, text="Start Monitor", command=self.start_monitor).pack(side="right")
        ttk.Button(controls, text="Stop", command=self.stop_monitor).pack(side="right", padx=(0, 8))

        ttk.Label(parent, textvariable=self.summary_text, style="Summary.TLabel").pack(fill="x", pady=(12, 8))

        columns = ("nurse", "document", "expires", "days", "bucket", "due", "row")
        self.table = ttk.Treeview(parent, columns=columns, show="headings", height=18)
        headings = {
            "nurse": "Nurse",
            "document": "Document",
            "expires": "Expires",
            "days": "Days",
            "bucket": "Status",
            "due": "Due",
            "row": "Excel Row",
        }
        widths = {"nurse": 220, "document": 170, "expires": 100, "days": 70, "bucket": 150, "due": 70, "row": 80}
        for column in columns:
            self.table.heading(column, text=headings[column])
            self.table.column(column, width=widths[column], anchor="w")
        self.table.pack(fill="both", expand=True)

        log_frame = ttk.LabelFrame(parent, text="Log", padding=8)
        log_frame.pack(fill="x", pady=(10, 0))
        self.log_box = ttk.Label(log_frame, text="", anchor="w", justify="left")
        self.log_box.pack(fill="x")

    def build_settings_tab(self, parent: ttk.Frame) -> None:
        parent.columnconfigure(1, weight=1)

        row = 0
        ttk.Label(parent, text="Excel file").grid(row=row, column=0, sticky="w", pady=5)
        ttk.Entry(parent, textvariable=self.vars["excel_path"]).grid(row=row, column=1, sticky="ew", pady=5, padx=8)
        ttk.Button(parent, text="Browse", command=self.browse_excel).grid(row=row, column=2, pady=5)

        row += 1
        ttk.Label(parent, text="Worksheet name").grid(row=row, column=0, sticky="w", pady=5)
        ttk.Entry(parent, textvariable=self.vars["sheet_name"]).grid(row=row, column=1, sticky="ew", pady=5, padx=8)
        ttk.Label(parent, text="Blank uses the first sheet").grid(row=row, column=2, sticky="w")

        row += 1
        ttk.Label(parent, text="HR copy emails").grid(row=row, column=0, sticky="nw", pady=5)
        self.copy_emails = ttk.Entry(parent, textvariable=self.vars["hr_copy_emails"])
        self.copy_emails.grid(row=row, column=1, sticky="ew", pady=5, padx=8)
        ttk.Label(parent, text="Separate with comma").grid(row=row, column=2, sticky="w")

        row += 1
        ttk.Separator(parent).grid(row=row, column=0, columnspan=3, sticky="ew", pady=14)

        for label, key in [
            ("SMTP host", "smtp_host"),
            ("SMTP port", "smtp_port"),
            ("SMTP username", "smtp_username"),
            ("SMTP password", "smtp_password"),
            ("From email", "smtp_from"),
            ("Subject prefix", "subject_prefix"),
        ]:
            row += 1
            ttk.Label(parent, text=label).grid(row=row, column=0, sticky="w", pady=5)
            show = "*" if key == "smtp_password" else ""
            ttk.Entry(parent, textvariable=self.vars[key], show=show).grid(row=row, column=1, sticky="ew", pady=5, padx=8)

        row += 1
        ttk.Checkbutton(parent, text="Use STARTTLS", variable=self.smtp_use_tls).grid(row=row, column=1, sticky="w", pady=5)

        row += 1
        ttk.Checkbutton(parent, text="Enable hourly monitor when app is open", variable=self.monitor_enabled).grid(row=row, column=1, sticky="w", pady=5)

        row += 1
        buttons = ttk.Frame(parent)
        buttons.grid(row=row, column=1, sticky="w", pady=16, padx=8)
        ttk.Button(buttons, text="Save Settings", command=self.save_current_settings).pack(side="left")
        ttk.Button(buttons, text="Send Test Email", command=self.send_test_email).pack(side="left", padx=(8, 0))

        row += 1
        note = (
            "Recommended Excel headers: Nurse Name, Email, Email to SMS, Document Type, Expiration Date. "
            "Email-to-SMS is sent through carrier gateway addresses."
        )
        ttk.Label(parent, text=note, wraplength=780, foreground="#475569").grid(row=row, column=0, columnspan=3, sticky="w", pady=10)

    def browse_excel(self) -> None:
        filename = filedialog.askopenfilename(
            title="Choose Excel file",
            filetypes=[("Excel files", "*.xlsx *.xlsm *.csv"), ("All files", "*.*")],
        )
        if filename:
            self.vars["excel_path"].set(filename)

    def current_settings(self) -> dict:
        settings = dict(DEFAULT_SETTINGS)
        for key, variable in self.vars.items():
            settings[key] = variable.get()
        settings["smtp_use_tls"] = self.smtp_use_tls.get()
        settings["monitor_enabled"] = self.monitor_enabled.get()
        return settings

    def save_current_settings(self) -> None:
        self.settings = self.current_settings()
        save_settings(self.settings)
        self.status("Settings saved.")
        if self.monitor_enabled.get():
            self.start_monitor()
        self.refresh_preview()

    def refresh_preview(self) -> None:
        self.settings = self.current_settings()
        try:
            alerts, warnings = read_alerts(self.settings)
            for item in self.table.get_children():
                self.table.delete(item)
            for alert in alerts:
                self.table.insert(
                    "",
                    "end",
                    values=(
                        alert.nurse_name,
                        alert.document_name,
                        alert.expires_on.isoformat(),
                        alert.days_until,
                        BUCKET_RULES[alert.bucket]["label"],
                        "Yes" if alert.due_now else "No",
                        alert.row_number,
                    ),
                )
            expired = sum(1 for alert in alerts if alert.bucket == "expired")
            due = sum(1 for alert in alerts if alert.due_now)
            self.summary_text.set(f"Found {len(alerts)} expired or upcoming records. {expired} expired. {due} due now.")
            self.log_box.configure(text="\n".join(warnings[:4]) if warnings else "No warnings.")
            self.status("Preview refreshed.")
        except Exception as exc:
            self.status("Preview failed.")
            self.log_box.configure(text=str(exc))
            log(f"Preview failed: {exc}")

    def run_send(self, force: bool) -> None:
        if self.worker_running:
            messagebox.showinfo("Still running", "A send job is already running.")
            return
        self.save_current_settings()
        thread = threading.Thread(target=self._send_worker, args=(force,), daemon=True)
        thread.start()

    def _send_worker(self, force: bool) -> None:
        self.worker_running = True
        self.root.after(0, lambda: self.status("Sending messages..."))
        try:
            result = send_due_messages(self.settings, force=force)
            text = f"Scanned {result['scanned']}. Sent {result['sent']}. Skipped {result['skipped']}."
            if result["warnings"]:
                text += "\n" + "\n".join(result["warnings"][:4])
            log(text.replace("\n", " | "))
            self.root.after(0, lambda: self.log_box.configure(text=text))
            self.root.after(0, self.refresh_preview)
        except Exception as exc:
            log(f"Send failed: {exc}")
            self.root.after(0, lambda: self.log_box.configure(text=f"Send failed: {exc}"))
            self.root.after(0, lambda: self.status("Send failed."))
        finally:
            self.worker_running = False

    def send_test_email(self) -> None:
        self.save_current_settings()
        recipients = split_emails(self.settings.get("hr_copy_emails", ""))
        if not recipients:
            messagebox.showwarning("No recipient", "Add at least one HR copy email first.")
            return
        try:
            send_email(
                self.settings,
                recipients[0],
                "Nurse License Monitor test",
                "This is a test message from the Nurse License Monitor desktop app.",
            )
            self.status("Test email sent.")
            messagebox.showinfo("Sent", f"Test email sent to {recipients[0]}.")
        except Exception as exc:
            log(f"Test email failed: {exc}")
            messagebox.showerror("Test failed", str(exc))

    def start_monitor(self) -> None:
        self.monitor_running = True
        self.monitor_enabled.set(True)
        self.status("Monitor running. Checks every hour.")
        self.root.after(1000, self.monitor_tick)

    def stop_monitor(self) -> None:
        self.monitor_running = False
        self.monitor_enabled.set(False)
        self.status("Monitor stopped.")

    def monitor_tick(self) -> None:
        if not self.monitor_running:
            return
        if not self.worker_running:
            thread = threading.Thread(target=self._send_worker, args=(False,), daemon=True)
            thread.start()
        self.root.after(60 * 60 * 1000, self.monitor_tick)

    def status(self, text: str) -> None:
        self.status_text.set(text)

    def run(self) -> None:
        self.root.mainloop()


if __name__ == "__main__":
    try:
        NurseLicenseMonitorApp().run()
    except Exception as error:
        log(f"Application crashed: {error}")
        raise
